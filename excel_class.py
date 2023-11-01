import openpyxl
import xlsxwriter
import datetime
import sys
import re


class excel_hours():
    def __init__(self, Path_to_file: str) -> None:
        self.months = {1: 'January', 2: 'February', 3: 'March', 4: 'April', 5: 'May', 6: 'June', 7: 'July',
                       8: 'August', 9: 'September', 10: 'October', 11: 'November', 12: 'December'}
        self.path = Path_to_file

    def new_sheet_and_fill(self, excel_file, worksheet_name, current_month):
        """
        Method to create new excel worksheet, create table inside and customize it

        Args:
            excel_file (str): Variable for name of file
            worksheet_name (str): Variable for worksheet name
            current_month (int): key in the dictionary so that extract value(month name) from dictionary and insert into table
        """
        def make_dates_for_respective_months(row_number):
            '''fill dates and mark weekends'''
            if (current_month == 2) and (row_number - 4 <= 28):
                date = datetime.date(
                    year=datetime.datetime.now().year, month=current_month, day=row_number-4)
                if date.weekday() in (5, 6):
                    for col in range(4, 7, 2):
                        worksheet_name.merge_range(row_number, col, row_number, col+1, None, excel_file.add_format(
                            {'bg_color': 'red', 'align': 'center'}))
                else:
                    for col in range(4, 7, 2):
                        worksheet_name.merge_range(
                            row_number, col, row_number, col+1, None, cell_border)

                worksheet_name.merge_range(
                    row_number, 2, row_number, 3, date.strftime('%d/%m/%Y'), cell_border)
            elif (current_month % 2 == 0 and current_month != 2 and current_month < 8) and (row_number - 4 <= 30)\
                    or (current_month % 2 != 0 and current_month != 2 and current_month >= 8) and (row_number - 4 <= 30):
                date = datetime.date(
                    year=datetime.datetime.now().year, month=current_month, day=row_number-4)
                if date.weekday() in (5, 6):
                    for col in range(4, 7, 2):
                        worksheet_name.merge_range(row_number, col, row_number, col+1, None, excel_file.add_format(
                            {'bg_color': 'red', 'align': 'center'}))
                else:
                    for col in range(4, 7, 2):
                        worksheet_name.merge_range(
                            row_number, col, row_number, col+1, None, cell_border)

                worksheet_name.merge_range(
                    row_number, 2, row_number, 3, date.strftime('%d/%m/%Y'), cell_border)
            else:
                if (current_month % 2 != 0 and current_month < 8) or (current_month % 2 == 0 and current_month >= 8):
                    date = datetime.date(
                        year=datetime.datetime.now().year, month=current_month, day=row_number-4)
                    if date.weekday() in (5, 6):
                        for col in range(4, 7, 2):
                            worksheet_name.merge_range(row_number, col, row_number, col+1, None, excel_file.add_format(
                                {'bg_color': 'red', 'align': 'center'}))
                    else:
                        for col in range(4, 7, 2):
                            worksheet_name.merge_range(
                                row_number, col, row_number, col+1, None, cell_border)

                    worksheet_name.merge_range(
                        row_number, 2, row_number, 3, date.strftime('%d/%m/%Y'), cell_border)

        def make_border_for_cells(len_month=36):
            for c in range(2, 8):
                for r in range(3, len_month):
                    worksheet_name.write(r, c, None, cell_border)

        '''main triggering for fill and stylize table'''

        cell_border = excel_file.add_format(
            {'border': 1, 'align': 'center', 'text_wrap': True})
        fst_cell = excel_file.add_format(
            {'border': 1, 'align': 'center', 'bg_color': '#D3D3D3'})

        if current_month == 2:
            make_border_for_cells(33)
        elif current_month % 2 == 0:
            if current_month > 7:
                make_border_for_cells()
            else:
                make_border_for_cells(35)
        else:
            if current_month > 7:
                make_border_for_cells(35)
            else:
                make_border_for_cells()

        worksheet_name.merge_range(
            2, 2, 2, 10, 'Table with work hours', fst_cell)
        worksheet_name.merge_range(
            3, 4, 4, 5, 'From hour to hour', cell_border)
        worksheet_name.merge_range(
            3, 6, 4, 7, 'Amount of hours for the day', cell_border)
        worksheet_name.merge_range(
            3, 8, 4, 10, 'Total hours in a whole month (not included weekends)', cell_border)
        worksheet_name.merge_range(
            5, 8, 6, 10, None, cell_border)

        # stylize and fill table + mark when is weekend
        for r in range(3, 36):
            if excel_file.get_worksheet_by_name(self.months[current_month][:3]):
                if r == 3:
                    worksheet_name.merge_range(
                        r, 2, r, 3, 'Month/Date', cell_border)
                elif r == 4:
                    worksheet_name.merge_range(
                        r, 2, r, 3, self.months[current_month], cell_border)
                else:
                    make_dates_for_respective_months(r)

    def user_action(self, choice):

        wb = openpyxl.load_workbook(filename=self.path)

        def insert_into_excel(particular_date):

            if self.months[int(particular_date[3:5])][0:3] in wb.sheetnames:
                ws_current = wb[self.months[int(particular_date[3:5])][0:3]]
                for row in ws_current.iter_rows(min_row=6, min_col=3, max_col=4, max_row=36):
                    for cell in row:
                        if cell.value == particular_date+f'/{datetime.datetime.now().year}':
                            add_h_range = ws_current.cell(
                                row=cell.row, column=cell.column+2)
                            add_amount_h = ws_current.cell(
                                row=cell.row, column=cell.column+4)
                            current_value = add_h_range.value
                            if current_value:
                                add_h_range.value = current_value +'&'+ respond_hours_range
                            else:
                                add_h_range.value = respond_hours_range
                            if respond_hours_range != "-":
                                # calculation of hours difference
                                if '&' not in add_h_range.value:
                                    diff = (datetime.datetime.strptime(respond_hours_range[6:11], "%H.%M") -
                                            datetime.timedelta(hours=int(respond_hours_range[0:2]), minutes=int(respond_hours_range[3:5]))).strftime("%H.%M")
                                    diff = float(diff)  
                                else:
                                    num_of_appearances = add_h_range.value.count('&')
                                    diff = []
                                    # idx_spread = add_h_range.value.index('&')
                                    if num_of_appearances >= 1:
                                        splitted = (current_value +'&'+ respond_hours_range).split('&')
                                        for noa in range(num_of_appearances+1):
                                            diff.append(float(((datetime.datetime.strptime(splitted[noa][6:11], "%H.%M") -
                                                datetime.timedelta(hours=int(splitted[noa][0:2]), minutes=int(splitted[noa][3:5]))).strftime("%H.%M")))) 
                                        diff = sum(diff)           
                                
                                after_decimal_point = round(
                                    (diff - int(diff))/0.6, 2)
                                # add amount of hours
                                add_amount_h.value = int(
                                    diff) + after_decimal_point
                            else:
                                add_amount_h.value = 0
                            break

        def delete_inserted_hours_from_excel(relevant_date):
            if self.months[int(relevant_date[3:5])][0:3] in wb.sheetnames:
                ws_current = wb[self.months[int(relevant_date[3:5])][0:3]]
                for col in ws_current.iter_cols(min_row=6, max_row=36, min_col=3, max_col=4):
                    for cell in col:
                        if cell.value == relevant_date+f'/{datetime.datetime.now().year}' and cell.fill.start_color.index == '00000000':
                            if ws_current[f'G{cell.row}'].value != None:
                                ws_current['I6'] = ws_current['I6'].value - \
                                    ws_current[f'G{cell.row}'].value
                                ws_current[f'E{cell.row}'].value = None
                                ws_current[f'G{cell.row}'].value = None
                                try:
                                    wb.save(self.path)
                                    return print('\nRemoved')
                                except PermissionError:
                                    print('File is opened! Firstly, close it.')
                                    sys.exit(0)
                            else:
                                return print(
                                    "For this particular date, hours haven't been inserted yet")
                        if cell.row == 34 and int(relevant_date[3:5]) == 'February':
                            break
                        if cell.row == 36 and int(relevant_date[3:5]) in {'April', 'June', 'September', 'November'}:
                            break
                    break

        def sum_total_hours_in_a_month():
            for s in self.months.values():
                sum = 0
                ws_current = wb[s[0:3]]
                for col in ws_current.iter_cols(min_row=6, min_col=7, max_col=8, max_row=36):
                    for cell in col:
                        if cell.row == 34 and s == 'February':
                            break
                        if cell.row == 36 and s in {'April', 'June', 'September', 'November'}:
                            break
                        if cell.fill.start_color.index == '00000000':
                            try:
                                sum += cell.value
                            except TypeError:
                                sum += 0
                    break
                ws_current['I6'] = sum
            try:
                wb.save(self.path)
            except PermissionError:
                print('File is opened! Firstly, close it.')
                sys.exit(0)

        if choice == '1':

            respond_date = input(
                'Please enter what date you want fill: Custom date or today\'s date (input format: custom/today): ')
            if respond_date == 'custom' or respond_date == 'today':
                respond_hours_range = input(
                    '''
Please enter start hour when you began your work and end hour when you end it(exemplary input format: 07.00-09.30)
or (-) when you didn't work at all: 
                    ''')
            else:
                raise Exception('Bad input format')
            # regular expression to check if input contains proper format
            if re.match('^[0-9]{2}.[0-9]{2}-[0-9]{2}.[0-9]{2}$', respond_hours_range) or respond_hours_range == "-":
                if respond_date == "custom":
                    respond_particular_date = input(
                        'Please enter particular day and month to whom you want insert hours (dd/mm): ')
                    insert_into_excel(respond_particular_date)
                    sum_total_hours_in_a_month()
                else:
                    date_today = datetime.date.today().strftime('%d/%m')
                    insert_into_excel(date_today)
                    sum_total_hours_in_a_month()
            else:
                raise Exception('Bad input format')
        else:
            data_to_remove = input(
                'Type day and month for whom you want to remove hours (dd/mm): ')
            # regular expression to check if input contains proper format
            if re.match('^[0-9]{2}/[0-9]{2}$', data_to_remove):

                delete_inserted_hours_from_excel(data_to_remove)
            else:
                raise Exception('Bad input format')

    def create_excel(self):
        new_excel_file = xlsxwriter.Workbook(
            self.path)

        for k, v in self.months.items():
            Work_sheet = new_excel_file.add_worksheet(f'{v[:3]}')
            self.new_sheet_and_fill(new_excel_file, Work_sheet, k)
        print("Excel file didn't exist, it was just created.")
        new_excel_file.close()
